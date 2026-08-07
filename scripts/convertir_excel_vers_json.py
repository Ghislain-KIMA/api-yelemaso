"""
Convertit un classeur Excel MULTI-FEUILLES (Region / Province / Commune /
Mairie, une feuille par table) en fixture JSON Django.

Corrige deux pièges classiques de la conversion Excel :
1. Ne jamais utiliser la "feuille active" du classeur -- chaque feuille
   est ouverte par son NOM exact, pour être certain de lire la bonne.
2. Les cellules fusionnées ne stockent la valeur que dans la cellule en
   haut à gauche de la fusion -- ce script "démerge" chaque plage
   fusionnée en propageant cette valeur à toutes les cellules qu'elle
   couvre, pour éviter les colonnes qui semblent vides à tort.

Usage :
    python convertir_excel_vers_json.py source.xlsx sortie.json

Adapte NOMS_FEUILLES et COLONNES ci-dessous à ton fichier réel avant de
lancer le script -- lance d'abord avec --lister pour voir les feuilles
et les en-têtes détectés, sans rien convertir :
    python convertir_excel_vers_json.py source.xlsx --lister
"""
import json
import sys

import openpyxl

# Nom EXACT de chaque feuille dans le classeur (sensible à la casse et
# aux espaces -- vérifie avec --lister si la conversion échoue).
NOMS_FEUILLES = {
    "region": "Region",
    "province": "Province",
    "commune": "Commune",
}

# Nom des colonnes attendues sur chaque feuille (en-tête, 1ère ligne).
COLONNES = {
    "region": {"nom": "nom_region"},
    "province": {"nom": "nom_province", "region": "nom_region"},
    "commune": {"nom": "nom_commune", "province": "nom_province", "type_commune": "type_commune"},
}


def demerger_feuille(feuille):
    """
    Propage la valeur de chaque cellule fusionnée (stockée uniquement en
    haut à gauche) vers toutes les cellules de la plage fusionnée, pour
    que la lecture ligne par ligne ne rencontre plus de "trous".
    """
    for plage in list(feuille.merged_cells.ranges):
        valeur = feuille.cell(row=plage.min_row, column=plage.min_col).value
        feuille.unmerge_cells(str(plage))
        for ligne in range(plage.min_row, plage.max_row + 1):
            for colonne in range(plage.min_col, plage.max_col + 1):
                feuille.cell(row=ligne, column=colonne).value = valeur


def lister(chemin_excel: str):
    classeur = openpyxl.load_workbook(chemin_excel)
    print("Feuilles trouvées dans le classeur :")
    for nom in classeur.sheetnames:
        feuille = classeur[nom]
        entetes = [c.value for c in feuille[1]]
        print(f"  - '{nom}'  (en-têtes : {entetes})")


def lire_feuille(classeur, cle: str):
    nom_feuille = NOMS_FEUILLES[cle]
    if nom_feuille not in classeur.sheetnames:
        raise SystemExit(
            f"Feuille '{nom_feuille}' introuvable. Feuilles disponibles : "
            f"{classeur.sheetnames}. Ajuste NOMS_FEUILLES dans le script, "
            f"ou lance avec --lister pour voir les noms exacts."
        )
    feuille = classeur[nom_feuille]
    demerger_feuille(feuille)

    entetes = [cell.value for cell in feuille[1]]
    colonnes_attendues = COLONNES[cle]
    idx = {}
    for champ, nom_colonne in colonnes_attendues.items():
        if nom_colonne not in entetes:
            raise SystemExit(
                f"Colonne '{nom_colonne}' introuvable sur la feuille "
                f"'{nom_feuille}'. En-têtes trouvés : {entetes}"
            )
        idx[champ] = entetes.index(nom_colonne)

    lignes = []
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in ligne):
            continue  # ligne complètement vide, ignorée
        lignes.append({champ: ligne[i] for champ, i in idx.items()})
    return lignes


def convertir(chemin_excel: str, chemin_json: str):
    classeur = openpyxl.load_workbook(chemin_excel)

    lignes_region = lire_feuille(classeur, "region")
    lignes_province = lire_feuille(classeur, "province")
    lignes_commune = lire_feuille(classeur, "commune")

    objets = []
    regions = {}     # nom -> pk
    provinces = {}   # nom -> pk

    for ligne in lignes_region:
        nom = ligne["nom"]
        if not nom or nom in regions:
            continue
        regions[nom] = len(regions) + 1
        objets.append({
            "model": "administration.region",
            "pk": regions[nom],
            "fields": {"nom_region": nom},
        })

    for ligne in lignes_province:
        nom = ligne["nom"]
        nom_region = ligne["region"]
        if not nom or nom in provinces:
            continue
        if nom_region not in regions:
            print(f"AVERTISSEMENT : région '{nom_region}' inconnue pour la "
                  f"province '{nom}' -- ligne ignorée.")
            continue
        provinces[nom] = len(provinces) + 1
        objets.append({
            "model": "administration.province",
            "pk": provinces[nom],
            "fields": {"nom_province": nom, "region": regions[nom_region]},
        })

    pk_commune = 0
    for ligne in lignes_commune:
        nom = ligne["nom"]
        nom_province = ligne["province"]
        if not nom:
            continue
        if nom_province not in provinces:
            print(f"AVERTISSEMENT : province '{nom_province}' inconnue pour "
                  f"la commune '{nom}' -- ligne ignorée.")
            continue
        pk_commune += 1
        objets.append({
            "model": "administration.commune",
            "pk": pk_commune,
            "fields": {
                "nom_commune": nom,
                "province": provinces[nom_province],
                # type_commune (nullable) : rempli seulement si présent
                # dans le fichier source, sinon laissé à null pour être
                # complété plus tard.
                "type_commune": ligne.get("type_commune") or None,
            },
        })

    with open(chemin_json, "w", encoding="utf-8") as f:
        json.dump(objets, f, ensure_ascii=False, indent=2)

    print(f"{len(regions)} régions, {len(provinces)} provinces, "
          f"{pk_commune} communes -> {chemin_json}")


if __name__ == "__main__":
    if len(sys.argv) == 3 and sys.argv[2] == "--lister":
        lister(sys.argv[1])
    elif len(sys.argv) == 3:
        convertir(sys.argv[1], sys.argv[2])
    else:
        print("Usage:")
        print("  python convertir_excel_vers_json.py source.xlsx --lister")
        print("  python convertir_excel_vers_json.py source.xlsx sortie.json")
        sys.exit(1)
