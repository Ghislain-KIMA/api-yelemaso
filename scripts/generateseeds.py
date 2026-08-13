import openpyxl

try:
    classeur = openpyxl.load_workbook("administratio.xlsx", read_only=True)
    for nom in classeur.sheetnames:
        feuille = classeur[nom]
        entetes = [c.value if c.value is not None else "" for c in feuille[1]] if feuille.max_row >= 1 else []
        print(f"  - {nom} (en-têtes : {entetes})")
except FileNotFoundError:
    print("Le fichier 'administration.xlsx' est introuvable.")
