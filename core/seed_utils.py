"""
Utilitaires de seed, partagés par les commandes seed_administration, 
seed_culture et seed_gestion.

Lit directement les fichiers .xlsx (pas de CSV intermédiaire -- évite les
pièges classiques de l'export Excel -> CSV : perte de mise en forme, export de
la "feuille active" plutôt que la bonne feuille, cellules fusionnées qui 
semblent vides).

Toujours idempotent (get_or_create) : relancer une commande de seed plusieurs
fois ne crée jamais de doublon.
"""



import glob
import os
import openpyxl



def trouver_fichiers_source(dossier_source: str) -> list[str]:
    """Liste tous les .xlsx d'un dossier source. Vide si le dossier n'existe pas."""
    if not os.path.isdir(dossier_source):
        return []
    return sorted(glob.glob(os.path.join(dossier_source, "*.xlsx")))


def ouvrir_classeurs(chemins_excel: list[str]) -> dict:
    """Ouvre plusieurs fichiers Excel -> {chemin: classeur}."""
    return {chemin: openpyxl.load_workbook(chemin) for chemin in chemins_excel}


def demerger_feuille(feuille):
    """
    Une cellule fusionnée ne stocke sa valeur que dans la cellule en haut à gauche de la fusion -- cette fonction
    recopie cette valeur dans toutes les cellules de la plage, pour qu'une lecture ligne par ligne classique ne
    tombe pas sur des cases vides.
    """

    for plage in list(feuille.merged_cells.ranges):
        valeur = feuille.cell(row=plage.min_row, column=plage.min_col).value
        feuille.unmerge_cells(str(plage))
        for ligne in range(plage.min_row, plage.max_row + 1):
            for colonne in range(plage.min_col, plage.max_col + 1):
                feuille.cell(row=ligne, column=colonne).value = valeur


def trouver_feuille(classeurs: dict, nom_feuille: str, stdout=None):
    """
    Cherche une feuille par son NOM (jamais par "feuille active") dans
    l'ensemble des classeurs fournis. Retourne (chemin, feuille), ou
    (None, None) si absente de tous les fichiers.
    """
    trouvees = [
        (chemin, classeur[nom_feuille])
        for chemin, classeur in classeurs.items()
        if nom_feuille in classeur.sheetnames
    ]
    if not trouvees:
        return None, None
    if len(trouvees) > 1 and stdout:
        fichiers = ", ".join(c for c, _ in trouvees)
        stdout.write(f"  ! '{nom_feuille}' présente dans plusieurs fichiers "
                     f"({fichiers}) -- utilisation de {trouvees[0][0]}.")
    return trouvees[0]


def lire_lignes(feuille, colonnes: dict) -> list[dict]:
    """
    Lit une feuille et retourne ses données sous forme de liste de
    dictionnaires. `colonnes` associe un nom de champ à un nom de
    colonne Excel -- la position de chaque colonne est retrouvée par
    son NOM, jamais par un numéro fixe.
    """
    demerger_feuille(feuille)
    entetes = [cell.value for cell in feuille[1]]

    idx = {}
    for champ, nom_colonne in colonnes.items():
        if nom_colonne not in entetes:
            raise ValueError(
                f"Colonne '{nom_colonne}' introuvable. En-têtes trouvés : {entetes}"
            )
        idx[champ] = entetes.index(nom_colonne)

    lignes = []
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in ligne):
            continue
        lignes.append({champ: ligne[i] for champ, i in idx.items()})
    return lignes


def seed_labels(model, classeurs: dict, nom_feuille: str, colonne: str = "label", stdout=None) -> int:
    """
    Peuple un modèle simple (juste un champ `label`) à partir d'une
    feuille nommée `nom_feuille`, cherchée dans l'ensemble des
    classeurs fournis. Retourne le nombre d'objets réellement créés.
    """
    chemin, feuille = trouver_feuille(classeurs, nom_feuille, stdout=stdout)
    if feuille is None:
        if stdout:
            stdout.write(f"  ! Feuille '{nom_feuille}' absente de tous les "
                         f"fichiers source -- ignorée.")
        return 0

    lignes = lire_lignes(feuille, {"valeur": colonne})
    nb_crees = 0
    for ligne in lignes:
        valeur = ligne["valeur"]
        if not valeur:
            continue
        valeur = str(valeur).strip()
        _, cree = model.objects.get_or_create(**{colonne: valeur})
        if cree:
            nb_crees += 1
            if stdout:
                stdout.write(f"  + {model.__name__} créé : {valeur}")
    return nb_crees
